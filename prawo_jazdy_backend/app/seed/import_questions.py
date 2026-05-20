"""
Import pytań egzaminacyjnych z pliku XLSX Ministerstwa Infrastruktury.
Importuje obie arkusze: 'katalog' (zweryfikowane) i 'W trakcie weryfikacji'.
"""

import os

import openpyxl
from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession

from app.models.question import Question, QuestionType, QuestionScope


async def import_questions(db: AsyncSession, xlsx_path: str) -> dict:
    """Importuje pytania z pliku XLSX."""
    if not os.path.exists(xlsx_path):
        raise FileNotFoundError(f"Nie znaleziono pliku: {xlsx_path}")

    wb = openpyxl.load_workbook(xlsx_path, read_only=True)
    stats = {"katalog": 0, "weryfikacja": 0, "skipped": 0}

    for sheet_name, is_verified in [("katalog", True), ("W trakcie weryfikacji", False)]:
        if sheet_name not in wb.sheetnames:
            continue

        ws = wb[sheet_name]
        rows = list(ws.iter_rows(min_row=2, values_only=True))

        for row in rows:
            if not row or not row[1] or not row[2]:
                stats["skipped"] += 1
                continue

            question_number = str(row[1]).strip()
            question_text = str(row[2]).strip() if row[2] else ""
            answer_a = str(row[3]).strip() if row[3] else None
            answer_b = str(row[4]).strip() if row[4] else None
            answer_c = str(row[5]).strip() if row[5] else None
            correct_answer = str(row[6]).strip().upper() if row[6] else ""
            media_filename = str(row[7]).strip() if row[7] else None
            scope_str = str(row[8]).strip() if row[8] else "PODSTAWOWY"
            points_str = str(row[9]).strip() if row[9] else "1"
            categories = str(row[10]).strip() if row[10] else ""

            # English translations (columns 15-18, 0-indexed)
            question_text_en = str(row[15]).strip() if len(row) > 15 and row[15] else None
            answer_a_en = str(row[16]).strip() if len(row) > 16 and row[16] else None
            answer_b_en = str(row[17]).strip() if len(row) > 17 and row[17] else None
            answer_c_en = str(row[18]).strip() if len(row) > 18 and row[18] else None

            # Determine question type
            if correct_answer in ("T", "N"):
                question_type = QuestionType.TAK_NIE
            else:
                question_type = QuestionType.ABC

            # Parse scope
            if scope_str == "SPECJALISTYCZNY":
                scope = QuestionScope.SPECJALISTYCZNY
            else:
                scope = QuestionScope.PODSTAWOWY

            # Parse points
            try:
                points = int(points_str)
            except ValueError:
                points = 1

            # Clean empty strings to None
            if answer_a == "":
                answer_a = None
            if answer_b == "":
                answer_b = None
            if answer_c == "":
                answer_c = None
            if answer_a_en == "":
                answer_a_en = None
            if answer_b_en == "":
                answer_b_en = None
            if answer_c_en == "":
                answer_c_en = None
            if question_text_en == "":
                question_text_en = None
            if media_filename == "":
                media_filename = None

            # Check if question already exists
            result = await db.execute(
                select(Question).where(
                    Question.question_number == question_number,
                    Question.is_verified == is_verified,
                )
            )
            existing = result.scalar_one_or_none()
            if existing:
                stats["skipped"] += 1
                continue

            question = Question(
                question_number=question_number,
                question_text=question_text,
                answer_a=answer_a,
                answer_b=answer_b,
                answer_c=answer_c,
                correct_answer=correct_answer,
                question_type=question_type,
                media_filename=media_filename,
                scope=scope,
                points=points,
                categories=categories,
                is_verified=is_verified,
                question_text_en=question_text_en,
                answer_a_en=answer_a_en,
                answer_b_en=answer_b_en,
                answer_c_en=answer_c_en,
            )
            db.add(question)

            if is_verified:
                stats["katalog"] += 1
            else:
                stats["weryfikacja"] += 1

    await db.flush()
    wb.close()
    return stats
